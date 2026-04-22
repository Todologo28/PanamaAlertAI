"""Rutas web (mapa, dashboard) con JSON para AJAX."""
import logging
import math
import re
import csv
from datetime import datetime, timedelta
from io import BytesIO, StringIO

import requests
from flask import Blueprint, render_template, session, jsonify, request, current_app, send_file
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from ..security import login_required, list_security_events
from ..extensions import db
from sqlalchemy import text
from ..services.evidence import list_incident_evidence, save_incident_evidence
from ..services.news_ingest import (
    cleanup_expired_news_incidents,
    load_sources,
    load_sync_state,
    maybe_auto_sync_news,
    save_sources,
    sync_sources,
)
from ..services.preferences import load_preferences, save_preferences, should_notify
from ..services.mailer import send_email, send_geofence_alert_email, mail_is_configured
from ..services.trust import build_reporter_profiles, explain_analysis

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - fallback visible at runtime
    Workbook = None
    Font = None
    PatternFill = None
    get_column_letter = None

logger = logging.getLogger(__name__)

bp = Blueprint("main", __name__)
PANAMA_LAT_RANGE = (7.0, 9.9)
PANAMA_LNG_RANGE = (-83.2, -77.0)
CHAT_HISTORY_LIMIT = 8
MAX_CHAT_MESSAGE_LEN = 500
GENERIC_ZONE_LABELS = {"panama, panama", "ciudad de panama", "panama", "capital"}
EXPORT_SHEET_FILL = "0F172A"
EXPORT_SHEET_TEXT = "F8FAFC"
BI_DATASET_LABELS = {
    "overview": "Resumen",
    "incidents": "Incidentes",
    "daily": "Diario",
    "hotspots": "Hotspots",
    "users": "Usuarios",
    "dictionary": "DiccionarioBI",
}
BI_FEED_MAX_AGE_SECONDS = 60 * 60 * 12
SEVERITY_LABELS = {
    1: "Baja",
    2: "Media",
    3: "Alta",
    4: "Critica",
    5: "Extrema",
}


def _bi_feed_signer():
    return URLSafeTimedSerializer(current_app.config["SECRET_KEY"], salt="panamaalert-bi-feed")


def _generate_bi_feed_token(dataset):
    return _bi_feed_signer().dumps({"dataset": str(dataset or "").strip().lower()})


def _validate_bi_feed_token(token, dataset):
    try:
        payload = _bi_feed_signer().loads(token, max_age=BI_FEED_MAX_AGE_SECONDS)
    except SignatureExpired as exc:
        raise ValueError("El token BI expiro. Genera uno nuevo desde el dashboard.") from exc
    except BadSignature as exc:
        raise ValueError("El token BI no es valido.") from exc
    expected_dataset = str(dataset or "").strip().lower()
    if payload.get("dataset") != expected_dataset:
        raise ValueError("El token BI no corresponde al dataset solicitado.")
    return payload


def _clean_text(value, field_name, max_len, min_len=1):
    text_value = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text_value) < min_len or len(text_value) > max_len:
        raise ValueError(f"{field_name} invalido ({min_len}-{max_len} chars)")
    return text_value


def _validate_panama_coords(lat, lng):
    if not (PANAMA_LAT_RANGE[0] <= lat <= PANAMA_LAT_RANGE[1] and PANAMA_LNG_RANGE[0] <= lng <= PANAMA_LNG_RANGE[1]):
        raise ValueError("Ubicacion fuera del territorio admitido de Panama")
    return lat, lng


def _normalize_radius(radius_value, min_value=0.3, max_value=25):
    radius = float(radius_value)
    if not (min_value <= radius <= max_value):
        raise ValueError(f"Radio invalido ({min_value}-{max_value} km)")
    return round(radius, 2)


def _sanitize_chat_history(history, limit=CHAT_HISTORY_LIMIT):
    clean_items = []
    for item in history or []:
        if not isinstance(item, dict):
            continue
        role = (item.get("role") or "").strip().lower()
        content = re.sub(r"\s+", " ", str(item.get("content") or "")).strip()
        if role not in ("user", "assistant") or not content:
            continue
        clean_items.append({"role": role, "content": content[:MAX_CHAT_MESSAGE_LEN]})
    return clean_items[-limit:]


def _nominatim_headers():
    app_name = current_app.config.get("APP_NAME", "PanamaAlert") if current_app else "PanamaAlert"
    return {
        "User-Agent": f"{app_name}/1.0 (geo-search)",
        "Accept-Language": "es,en;q=0.8"
    }


def _geocode_place(query: str):
    query = (query or "").strip()
    if not query:
        raise ValueError("Lugar requerido")

    try:
        response = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={
                "q": query,
                "format": "jsonv2",
                "limit": 5,
                "addressdetails": 1,
                "countrycodes": "pa"
            },
            headers=_nominatim_headers(),
            timeout=8
        )
        response.raise_for_status()
        payload = response.json() or []
    except requests.RequestException as exc:
        logger.warning("Geocoding search failed: %s", exc)
        raise RuntimeError("No se pudo consultar ubicaciones en este momento")

    return [{
        "name": item.get("display_name") or query,
        "lat": float(item.get("lat")),
        "lng": float(item.get("lon"))
    } for item in payload if item.get("lat") and item.get("lon")]


def _reverse_geocode(lat: float, lng: float):
    try:
        response = requests.get(
            "https://nominatim.openstreetmap.org/reverse",
            params={
                "lat": lat,
                "lon": lng,
                "format": "jsonv2",
                "zoom": 17,
                "addressdetails": 1
            },
            headers=_nominatim_headers(),
            timeout=8
        )
        response.raise_for_status()
        payload = response.json() or {}
    except requests.RequestException as exc:
        logger.warning("Reverse geocoding failed: %s", exc)
        return None

    return payload.get("display_name")


def _extract_description_field(description, label):
    match = re.search(rf"{re.escape(label)}:\s*([^\n]+)", str(description or ""), flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def _incident_source_label(payload):
    description = str(payload.get("description") or "")
    match = re.search(r"\[Fuente externa:\s*([^\]]+)\]", description, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return payload.get("reporter_username") or payload.get("username") or "usuario"


def _tokenize_similarity_text(*values):
    tokens = set()
    for value in values:
        clean = re.sub(r"[^a-z0-9áéíóúñü ]+", " ", str(value or "").lower())
        tokens.update(part for part in clean.split() if len(part) >= 4)
    return tokens


def _distance_meters(lat1, lng1, lat2, lng2):
    if None in (lat1, lng1, lat2, lng2):
        return None
    mean_lat = math.radians((lat1 + lat2) / 2.0)
    dx = (lng2 - lng1) * 111320 * math.cos(mean_lat)
    dy = (lat2 - lat1) * 110540
    return math.sqrt(dx * dx + dy * dy)


def _build_incident_moderation_insights(results):
    normalized = []
    for item in results:
        try:
            lat = float(item.get("lat"))
            lng = float(item.get("lng"))
        except (TypeError, ValueError):
            lat, lng = None, None
        normalized.append((item, lat, lng))

    for item, lat, lng in normalized:
        zone_label = _extract_description_field(item.get("description"), "Zona objetivo")
        audience_label = _extract_description_field(item.get("description"), "Audiencia")
        source_label = _incident_source_label(item)
        zone_norm = re.sub(r"\s+", " ", zone_label.strip().lower())
        location_precision = "specific"
        if not zone_label:
            location_precision = "unknown"
        elif zone_norm in GENERIC_ZONE_LABELS:
            location_precision = "generic"

        title_tokens = _tokenize_similarity_text(item.get("title"), item.get("description"))
        duplicates = []
        for other, other_lat, other_lng in normalized:
            if other is item:
                continue
            if str(other.get("status")) == "dismissed":
                continue
            other_tokens = _tokenize_similarity_text(other.get("title"), other.get("description"))
            common = len(title_tokens & other_tokens)
            distance = _distance_meters(lat, lng, other_lat, other_lng)
            if common < 2 and (distance is None or distance > 240):
                continue
            duplicates.append({
                "incident_id": _incident_public_id(other),
                "title": other.get("title"),
                "status": other.get("status"),
                "distance_m": int(round(distance)) if distance is not None else None,
                "common_terms": common,
            })

        duplicates.sort(key=lambda row: ((row["distance_m"] if row["distance_m"] is not None else 999999), -row["common_terms"]))
        item["moderation_insights"] = {
            "source_label": source_label,
            "zone_label": zone_label or None,
            "audience_label": audience_label or None,
            "location_precision": location_precision,
            "possible_duplicates": duplicates[:3],
        }


def _safe_number(value, digits=2):
    try:
        return round(float(value), digits)
    except (TypeError, ValueError):
        return None


def _safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _serialize_export_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def _to_datetime(value):
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _day_name_es(value):
    names = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
    return names[value.weekday()] if value else ""


def _month_name_es(value):
    names = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    return names[value.month] if value else ""


def _source_type_label(row):
    reporter = str(row.get("reporter_username") or "").strip().lower()
    description = str(row.get("description") or "").lower()
    if reporter == "newsbot" or "[fuente externa:" in description:
        return "Bot / Fuente externa"
    return "Reporte ciudadano"


def _severity_band(value):
    sev = _safe_int(value)
    if sev >= 4:
        return "Alta criticidad"
    if sev == 3:
        return "Media criticidad"
    return "Baja criticidad"


def _build_bi_dictionary():
    return [
        {"dataset": "Incidentes", "field": "incident_id", "description": "Identificador unico del incidente"},
        {"dataset": "Incidentes", "field": "incident_date", "description": "Fecha del incidente en formato fecha"},
        {"dataset": "Incidentes", "field": "incident_hour", "description": "Hora numerica para analisis horario"},
        {"dataset": "Incidentes", "field": "severity_label", "description": "Etiqueta legible de severidad"},
        {"dataset": "Incidentes", "field": "source_type", "description": "Origen del dato: bot/fuente externa o reporte ciudadano"},
        {"dataset": "Diario", "field": "day_name", "description": "Nombre del dia para comparacion semanal"},
        {"dataset": "Diario", "field": "month_name", "description": "Nombre del mes para ejes y filtros"},
        {"dataset": "Hotspots", "field": "severity_band", "description": "Banda de criticidad agregada por celda"},
        {"dataset": "Usuarios", "field": "verification_rate", "description": "Porcentaje de reportes verificados por usuario"},
        {"dataset": "Resumen", "field": "metric/value", "description": "Pares KPI listos para cards o indicadores"},
    ]


def _transform_incident_export(row):
    created_at = _to_datetime(row.get("created_at"))
    verified_at = _to_datetime(row.get("verified_at"))
    severity = _safe_int(row.get("severity"))
    lat = _safe_number(row.get("lat"), 6)
    lng = _safe_number(row.get("lng"), 6)
    return {
        **dict(row),
        "incident_date": created_at.date().isoformat() if created_at else "",
        "incident_time": created_at.strftime("%H:%M:%S") if created_at else "",
        "incident_year": created_at.year if created_at else "",
        "incident_month": created_at.month if created_at else "",
        "incident_month_name": _month_name_es(created_at),
        "incident_day": created_at.day if created_at else "",
        "incident_hour": created_at.hour if created_at else "",
        "incident_day_name": _day_name_es(created_at),
        "verified_date": verified_at.date().isoformat() if verified_at else "",
        "severity_label": SEVERITY_LABELS.get(severity, "Sin clasificar"),
        "severity_band": _severity_band(severity),
        "is_verified": 1 if str(row.get("status") or "") == "verified" else 0,
        "is_pending": 1 if str(row.get("status") or "") == "pending" else 0,
        "is_dismissed": 1 if str(row.get("status") or "") == "dismissed" else 0,
        "geo_point": f"{lat},{lng}" if lat is not None and lng is not None else "",
        "source_type": _source_type_label(row),
    }


def _transform_daily_export(row):
    day_value = _to_datetime(row.get("day"))
    total = _safe_int(row.get("total_incidents"))
    verified = _safe_int(row.get("verified_count"))
    return {
        **dict(row),
        "year": day_value.year if day_value else "",
        "month": day_value.month if day_value else "",
        "month_name": _month_name_es(day_value),
        "day_of_month": day_value.day if day_value else "",
        "day_name": _day_name_es(day_value),
        "verification_rate": round((verified / total) * 100, 2) if total else 0,
        "severity_band": _severity_band(row.get("avg_severity")),
    }


def _transform_hotspot_export(row):
    last_incident_at = _to_datetime(row.get("last_incident_at"))
    return {
        **dict(row),
        "severity_band": _severity_band(row.get("avg_severity")),
        "last_incident_date": last_incident_at.date().isoformat() if last_incident_at else "",
    }


def _transform_user_export(row):
    incidents_reported = _safe_int(row.get("incidents_reported"))
    verified_reports = _safe_int(row.get("verified_reports"))
    joined_at = _to_datetime(row.get("joined_at"))
    return {
        **dict(row),
        "verification_rate": round((verified_reports / incidents_reported) * 100, 2) if incidents_reported else 0,
        "joined_year": joined_at.year if joined_at else "",
        "joined_month": joined_at.month if joined_at else "",
        "joined_month_name": _month_name_es(joined_at),
        "has_paid_plan": 0 if str(row.get("current_plan") or "").lower() in {"", "free", "none"} else 1,
    }


def _fetch_bi_export_payload():
    overview = db.session.execute(text("""
        SELECT
            COUNT(*) AS total_incidents,
            SUM(CASE WHEN status='verified' THEN 1 ELSE 0 END) AS verified_incidents,
            SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) AS pending_incidents,
            SUM(CASE WHEN status='dismissed' THEN 1 ELSE 0 END) AS dismissed_incidents,
            AVG(severity) AS avg_severity,
            MAX(created_at) AS last_incident_at
        FROM incidents
    """)).mappings().first() or {}

    incidents = db.session.execute(text("""
        SELECT
            incident_id,
            title,
            description,
            category_name,
            province_name,
            district_name,
            lat,
            lng,
            severity,
            status,
            reporter_username,
            created_at,
            verified_at,
            score,
            comments_count
        FROM v_incidents_full
        ORDER BY created_at DESC
        LIMIT 2000
    """)).mappings().all()

    daily = db.session.execute(text("""
        SELECT
            day,
            province,
            district,
            category,
            total_incidents,
            verified_count,
            resolved_count,
            avg_severity,
            max_severity
        FROM v_incidents_daily_stats
        ORDER BY day DESC, province, district, category
        LIMIT 2000
    """)).mappings().all()

    hotspots = db.session.execute(text("""
        SELECT
            lat_cell,
            lng_cell,
            incidents_30d,
            avg_severity,
            last_incident_at
        FROM v_hotspots
        ORDER BY incidents_30d DESC, avg_severity DESC
        LIMIT 500
    """)).mappings().all()

    users = db.session.execute(text("""
        SELECT
            user_id,
            username,
            full_name,
            email,
            role,
            current_plan,
            price_monthly_usd,
            incidents_reported,
            verified_reports,
            last_report_at,
            last_login_at,
            joined_at
        FROM v_user_activity
        ORDER BY incidents_reported DESC, verified_reports DESC
        LIMIT 1000
    """)).mappings().all()

    return {
        "generated_at": datetime.utcnow(),
        "overview": {
            "total_incidents": _safe_int(overview.get("total_incidents")),
            "verified_incidents": _safe_int(overview.get("verified_incidents")),
            "pending_incidents": _safe_int(overview.get("pending_incidents")),
            "dismissed_incidents": _safe_int(overview.get("dismissed_incidents")),
            "avg_severity": _safe_number(overview.get("avg_severity")),
            "last_incident_at": overview.get("last_incident_at"),
        },
        "incidents": [_transform_incident_export(dict(row)) for row in incidents],
        "daily": [_transform_daily_export(dict(row)) for row in daily],
        "hotspots": [_transform_hotspot_export(dict(row)) for row in hotspots],
        "users": [_transform_user_export(dict(row)) for row in users],
        "dictionary": _build_bi_dictionary(),
    }


def _sheet_rows_from_payload(payload):
    overview = payload["overview"]
    return {
        BI_DATASET_LABELS["overview"]: [
            {"metric": "Generado", "value": _serialize_export_value(payload["generated_at"])},
            {"metric": "Total incidentes", "value": overview["total_incidents"]},
            {"metric": "Verificados", "value": overview["verified_incidents"]},
            {"metric": "Pendientes", "value": overview["pending_incidents"]},
            {"metric": "Descartados", "value": overview["dismissed_incidents"]},
            {"metric": "Severidad promedio", "value": overview["avg_severity"]},
            {"metric": "Ultimo incidente", "value": _serialize_export_value(overview["last_incident_at"])},
        ],
        BI_DATASET_LABELS["incidents"]: payload["incidents"],
        BI_DATASET_LABELS["daily"]: payload["daily"],
        BI_DATASET_LABELS["hotspots"]: payload["hotspots"],
        BI_DATASET_LABELS["users"]: payload["users"],
        BI_DATASET_LABELS["dictionary"]: payload["dictionary"],
    }


def _csv_dataset_rows(payload, dataset):
    dataset = str(dataset or "").strip().lower()
    if dataset == "overview":
        return _sheet_rows_from_payload(payload)[BI_DATASET_LABELS["overview"]]
    if dataset == "incidents":
        return payload["incidents"]
    if dataset == "daily":
        return payload["daily"]
    if dataset == "hotspots":
        return payload["hotspots"]
    if dataset == "users":
        return payload["users"]
    if dataset == "dictionary":
        return payload["dictionary"]
    raise ValueError("Dataset BI no soportado")


def _build_csv_export(payload, dataset):
    rows = _csv_dataset_rows(payload, dataset)
    stream = StringIO()
    if rows:
        headers = list(rows[0].keys())
        writer = csv.DictWriter(stream, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: _serialize_export_value(row.get(header)) for header in headers})
    else:
        stream.write("sin_datos\n")
    data = BytesIO(stream.getvalue().encode("utf-8-sig"))
    data.seek(0)
    return data


def _build_bi_manifest(payload):
    generated_at = payload["generated_at"].strftime("%Y-%m-%d %H:%M:%S")
    return {
        "generated_at": generated_at,
        "recommended_connector": "Power BI / Excel unico o Power Query (.pq)",
        "datasets": [
            {
                "key": "incidents",
                "label": BI_DATASET_LABELS["incidents"],
                "route": "/api/bi/export.csv?dataset=incidents",
                "feed_route": "/api/bi/feed.csv?dataset=incidents&token=<token>",
                "grain": "1 fila por incidente",
                "recommended_visuals": ["Mapa", "Tabla detalle", "Segmentadores por estado y categoria"],
            },
            {
                "key": "daily",
                "label": BI_DATASET_LABELS["daily"],
                "route": "/api/bi/export.csv?dataset=daily",
                "feed_route": "/api/bi/feed.csv?dataset=daily&token=<token>",
                "grain": "1 fila por dia/provincia/distrito/categoria",
                "recommended_visuals": ["Series temporales", "Tendencia por provincia", "Stacked bars por categoria"],
            },
            {
                "key": "hotspots",
                "label": BI_DATASET_LABELS["hotspots"],
                "route": "/api/bi/export.csv?dataset=hotspots",
                "feed_route": "/api/bi/feed.csv?dataset=hotspots&token=<token>",
                "grain": "1 celda geografica agregada",
                "recommended_visuals": ["Mapa de calor", "Bubble map", "Ranking de celdas"],
            },
            {
                "key": "users",
                "label": BI_DATASET_LABELS["users"],
                "route": "/api/bi/export.csv?dataset=users",
                "feed_route": "/api/bi/feed.csv?dataset=users&token=<token>",
                "grain": "1 fila por usuario",
                "recommended_visuals": ["Productividad", "Retencion", "Planes actuales"],
            },
            {
                "key": "overview",
                "label": BI_DATASET_LABELS["overview"],
                "route": "/api/bi/export.csv?dataset=overview",
                "feed_route": "/api/bi/feed.csv?dataset=overview&token=<token>",
                "grain": "metric/value",
                "recommended_visuals": ["Cards KPI", "Indicadores ejecutivos"],
            },
            {
                "key": "dictionary",
                "label": BI_DATASET_LABELS["dictionary"],
                "route": "/api/bi/export.csv?dataset=dictionary",
                "feed_route": "/api/bi/feed.csv?dataset=dictionary&token=<token>",
                "grain": "1 fila por campo documentado",
                "recommended_visuals": ["Referencia de modelo", "Gobierno de datos"],
            },
        ],
    }


def _build_power_bi_guide_rows(base_url, payload):
    generated_at = _serialize_export_value(payload["generated_at"])
    return [
        {"step": 1, "action": "Archivo recomendado", "detail": "Usa este Excel como paquete unico para Power BI. Ya incluye resumen, detalle, hotspots, usuarios, diccionario y guia."},
        {"step": 2, "action": "Origen web alterno", "detail": f"Si prefieres conexion directa, usa el archivo .pq generado por la app con base {base_url}"},
        {"step": 3, "action": "Tabla principal", "detail": "Incidentes = detalle fila a fila. Usala para mapa, tabla y filtros principales."},
        {"step": 4, "action": "Tabla temporal", "detail": "Diario = tendencias por fecha, provincia, distrito y categoria."},
        {"step": 5, "action": "Tabla geografica", "detail": "Hotspots = mapa agregado para calor o burbujas."},
        {"step": 6, "action": "Tabla usuarios", "detail": "Usuarios = productividad, verificacion y planes."},
        {"step": 7, "action": "KPIs", "detail": "Resumen = tarjetas ejecutivas. No la relaciones con las demas."},
        {"step": 8, "action": "Documentacion", "detail": "DiccionarioBI = descripcion de campos para explicar el modelo en clase o anexos."},
        {"step": 9, "action": "Generado", "detail": f"Paquete BI generado: {generated_at}"},
    ]


def _build_power_bi_manifest_rows(payload, base_url):
    manifest = _build_bi_manifest(payload)
    rows = []
    for item in manifest["datasets"]:
        rows.append({
            "dataset_key": item["key"],
            "label": item["label"],
            "grain": item["grain"],
            "route": f"{base_url}{item['route']}",
            "recommended_visuals": ", ".join(item["recommended_visuals"]),
        })
    return rows


def _build_power_query_script(base_url):
    root = str(base_url or "").rstrip("/")
    token_map = {key: _generate_bi_feed_token(key) for key in BI_DATASET_LABELS.keys()}
    script = f"""let
    BaseUrl = "{root}",
    CsvFeed = (dataset as text, token as text) =>
        Csv.Document(
            Web.Contents(
                BaseUrl,
                [RelativePath="api/bi/feed.csv", Query=[dataset=dataset, token=token]]
            ),
            [Delimiter=",", Encoding=65001, QuoteStyle=QuoteStyle.Csv]
        ),
    Promote = (table as table) => Table.PromoteHeaders(table, [PromoteAllScalars=true]),
    Incidentes = Promote(CsvFeed("incidents", "{token_map["incidents"]}")),
    Diario = Promote(CsvFeed("daily", "{token_map["daily"]}")),
    Hotspots = Promote(CsvFeed("hotspots", "{token_map["hotspots"]}")),
    Usuarios = Promote(CsvFeed("users", "{token_map["users"]}")),
    Resumen = Promote(CsvFeed("overview", "{token_map["overview"]}")),
    DiccionarioBI = Promote(CsvFeed("dictionary", "{token_map["dictionary"]}"))
in
    [
        Incidentes = Incidentes,
        Diario = Diario,
        Hotspots = Hotspots,
        Usuarios = Usuarios,
        Resumen = Resumen,
        DiccionarioBI = DiccionarioBI
    ]"""
    return BytesIO(script.encode("utf-8"))


def _build_excel_export(payload, base_url):
    if Workbook is None:
        raise RuntimeError("La libreria openpyxl no esta disponible en este entorno")

    workbook = Workbook()
    first_sheet = True
    header_fill = PatternFill(fill_type="solid", fgColor=EXPORT_SHEET_FILL)
    header_font = Font(color=EXPORT_SHEET_TEXT, bold=True)
    sheets = _sheet_rows_from_payload(payload)
    sheets["ManifestBI"] = _build_power_bi_manifest_rows(payload, base_url)
    sheets["GuiaPowerBI"] = _build_power_bi_guide_rows(base_url, payload)

    for title, rows in sheets.items():
        sheet = workbook.active if first_sheet else workbook.create_sheet()
        sheet.title = title[:31]
        first_sheet = False

        if not rows:
            sheet.append(["Sin datos disponibles"])
            continue

        headers = list(rows[0].keys())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            sheet.append([_serialize_export_value(row.get(header)) for header in headers])

        for idx, header in enumerate(headers, start=1):
            values = [_serialize_export_value(header)]
            values.extend(_serialize_export_value(row.get(header)) for row in rows)
            width = max(len(str(value or "")) for value in values) + 2
            sheet.column_dimensions[get_column_letter(idx)].width = min(max(width, 12), 42)

        sheet.freeze_panes = "A2"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _build_dashboard_svg(payload):
    overview = payload["overview"]
    cards = [
        ("Total", overview["total_incidents"], "#2563eb"),
        ("Verificados", overview["verified_incidents"], "#059669"),
        ("Pendientes", overview["pending_incidents"], "#d97706"),
        ("Descartados", overview["dismissed_incidents"], "#dc2626"),
    ]
    top_categories = {}
    for row in payload["daily"]:
        category = row.get("category") or "Sin categoria"
        top_categories[category] = top_categories.get(category, 0) + _safe_int(row.get("total_incidents"))
    top_lines = sorted(top_categories.items(), key=lambda item: item[1], reverse=True)[:4]
    lines = []
    y = 242
    for idx, (name, total) in enumerate(top_lines, start=1):
        lines.append(
            f"<text x='54' y='{y}' font-size='18' font-weight='700' fill='#0f172a'>{idx}. {name}</text>"
            f"<text x='720' y='{y}' text-anchor='end' font-size='18' font-weight='700' fill='#334155'>{total}</text>"
        )
        y += 36

    card_svg = []
    start_x = 40
    for label, value, color in cards:
        card_svg.append(
            f"<rect x='{start_x}' y='88' width='170' height='94' rx='20' fill='#ffffff' stroke='#e2e8f0'/>"
            f"<text x='{start_x + 18}' y='120' font-size='16' font-weight='700' fill='#475569'>{label}</text>"
            f"<text x='{start_x + 18}' y='158' font-size='32' font-weight='800' fill='{color}'>{value}</text>"
        )
        start_x += 185

    svg = f"""<svg xmlns='http://www.w3.org/2000/svg' width='900' height='420' viewBox='0 0 900 420' role='img' aria-label='Resumen BI PanamaAlert'>
<rect width='900' height='420' fill='#f8fafc'/>
<rect x='24' y='24' width='852' height='372' rx='28' fill='url(#panel)' stroke='#dbeafe'/>
<defs>
  <linearGradient id='panel' x1='0%' y1='0%' x2='100%' y2='100%'>
    <stop offset='0%' stop-color='#eff6ff'/>
    <stop offset='100%' stop-color='#ffffff'/>
  </linearGradient>
</defs>
<text x='40' y='58' font-size='28' font-weight='800' fill='#0f172a'>PanamaAlert · Export ejecutivo BI</text>
<text x='40' y='80' font-size='14' fill='#475569'>Generado {payload["generated_at"].strftime("%Y-%m-%d %H:%M:%S")} · Severidad promedio {overview["avg_severity"] or 0}</text>
{''.join(card_svg)}
<text x='40' y='224' font-size='20' font-weight='800' fill='#0f172a'>Categorias con mas movimiento</text>
<line x1='40' y1='232' x2='860' y2='232' stroke='#cbd5e1'/>
{''.join(lines) if lines else "<text x='54' y='270' font-size='18' fill='#64748b'>Sin datos suficientes para el resumen.</text>"}
<text x='40' y='370' font-size='13' fill='#64748b'>Archivo pensado para evidencia visual, presentacion y anexos de Tableau/Power BI.</text>
</svg>"""
    return BytesIO(svg.encode("utf-8"))


# ============================================================================
# PÁGINAS HTML
# ============================================================================

@bp.route("/")
def public_map():
    return render_template("mapa.html", public_view=True)


@bp.route("/app")
@login_required
def home():
    return render_template("mapa.html", public_view=False)


@bp.route("/admin")
@login_required
def admin():
    """Panel de moderación — solo admins."""
    from ..models import User
    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() != "admin":
        return jsonify({"error": "No autorizado"}), 403
    return render_template("admin.html")


@bp.route("/profile")
@login_required
def profile():
    return render_template("profile.html")


@bp.route("/dashboard")
@login_required
def dashboard():
    ctx = _get_plan_context(session.get("user_id"))
    return render_template(
        "dashboard.html",
        assistant_mode=ctx["assistant_mode"],
        assistant_mode_label=ctx["assistant_mode_label"],
        plan_name=ctx["plan_name"],
        assistant_capabilities=ctx["assistant_capabilities"],
        assistant_sources=ctx["assistant_sources"],
        assistant_guardrails=ctx["assistant_guardrails"],
    )


@bp.route("/api/bi/export.xlsx", methods=["GET"])
@login_required
def api_bi_export_excel():
    try:
        payload = _fetch_bi_export_payload()
        stream = _build_excel_export(payload, request.url_root.rstrip("/"))
        stamp = payload["generated_at"].strftime("%Y%m%d_%H%M%S")
        return send_file(
            stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"panamaalert_bi_export_{stamp}.xlsx",
            max_age=0,
        )
    except Exception as exc:
        logger.exception("Excel BI export failed")
        return jsonify({"error": f"No se pudo generar el Excel BI: {exc}"}), 500


@bp.route("/api/bi/feed.csv", methods=["GET"])
def api_bi_feed_csv():
    dataset = request.args.get("dataset", "incidents", type=str)
    token = request.args.get("token", "", type=str)
    try:
        _validate_bi_feed_token(token, dataset)
        payload = _fetch_bi_export_payload()
        stream = _build_csv_export(payload, dataset)
        return send_file(
            stream,
            mimetype="text/csv; charset=utf-8",
            as_attachment=False,
            download_name=f"panamaalert_{dataset}.csv",
            max_age=0,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        logger.exception("CSV BI feed failed")
        return jsonify({"error": f"No se pudo generar el feed BI: {exc}"}), 500


@bp.route("/api/bi/export.svg", methods=["GET"])
@login_required
def api_bi_export_svg():
    try:
        payload = _fetch_bi_export_payload()
        stream = _build_dashboard_svg(payload)
        stamp = payload["generated_at"].strftime("%Y%m%d_%H%M%S")
        return send_file(
            stream,
            mimetype="image/svg+xml",
            as_attachment=True,
            download_name=f"panamaalert_dashboard_{stamp}.svg",
            max_age=0,
        )
    except Exception as exc:
        logger.exception("SVG BI export failed")
        return jsonify({"error": f"No se pudo generar el resumen SVG: {exc}"}), 500


@bp.route("/api/bi/export.csv", methods=["GET"])
@login_required
def api_bi_export_csv():
    dataset = request.args.get("dataset", "incidents", type=str)
    try:
        payload = _fetch_bi_export_payload()
        stream = _build_csv_export(payload, dataset)
        stamp = payload["generated_at"].strftime("%Y%m%d_%H%M%S")
        return send_file(
            stream,
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=f"panamaalert_{dataset}_{stamp}.csv",
            max_age=0,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        logger.exception("CSV BI export failed")
        return jsonify({"error": f"No se pudo generar el CSV BI: {exc}"}), 500


@bp.route("/api/bi/manifest", methods=["GET"])
@login_required
def api_bi_manifest():
    try:
        payload = _fetch_bi_export_payload()
        return jsonify(_build_bi_manifest(payload))
    except Exception as exc:
        logger.exception("BI manifest failed")
        return jsonify({"error": f"No se pudo generar el manifest BI: {exc}"}), 500


@bp.route("/api/bi/powerquery.pq", methods=["GET"])
@login_required
def api_bi_powerquery():
    try:
        stream = _build_power_query_script(request.url_root.rstrip("/"))
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return send_file(
            stream,
            mimetype="text/plain; charset=utf-8",
            as_attachment=True,
            download_name=f"panamaalert_powerbi_connector_{stamp}.pq",
            max_age=0,
        )
    except Exception as exc:
        logger.exception("Power Query BI export failed")
        return jsonify({"error": f"No se pudo generar el archivo Power Query: {exc}"}), 500


def _get_plan_context(user_id):
    """Obtiene el plan actual y capacidades visibles del asistente."""
    from ..models import Subscription, Plan

    sub = Subscription.query.filter_by(user_id=user_id).order_by(Subscription.started_at.desc()).first()
    plan = Plan.query.get(sub.plan_id) if sub else Plan.query.filter_by(name="Free").first()
    plan_name = plan.name if plan else "Free"
    assistant_mode = {
        "Free": "basic",
        "Pro": "advanced",
        "Enterprise": "enterprise"
    }.get(plan_name, "basic")
    assistant_mode_label = {
        "basic": "Asistente básico",
        "advanced": "Asistente avanzado",
        "enterprise": "Asistente con contexto completo"
    }.get(assistant_mode, "Asistente básico")

    capability_map = {
        "basic": [
            "Resume el riesgo actual de la zona y los incidentes cercanos.",
            "Responde preguntas frecuentes con contexto geografico real.",
            "Entrega recomendaciones breves para moverte con mas seguridad."
        ],
        "advanced": [
            "Explica patrones, tendencia y categorias dominantes por zona.",
            "Mantiene contexto reciente del chat para responder mejor.",
            "Prioriza acciones concretas segun el nivel de riesgo detectado."
        ],
        "enterprise": [
            "Amplia el contexto operativo para monitoreo y seguimiento continuo.",
            "Entrega respuestas mas completas para priorizar decisiones del equipo.",
            "Hace visibles los indicadores clave para una lectura mas ejecutiva."
        ]
    }
    assistant_sources = [
        "Incidentes cercanos dentro del radio seleccionado.",
        "Severidad promedio, volumen reciente y tendencia de la zona.",
        "Geocodificacion del lugar para aterrizar el contexto del usuario."
    ]
    assistant_guardrails = [
        "Cuando la IA no esta disponible, responde con reglas y estadisticas reales de la app.",
        "Las respuestas distinguen entre resumen de datos, tendencia y recomendaciones.",
        "No sustituye verificacion humana ni confirmacion de campo."
    ]

    return {
        "plan": plan,
        "subscription": sub,
        "plan_name": plan_name,
        "assistant_access": True,
        "assistant_mode": assistant_mode,
        "assistant_mode_label": assistant_mode_label,
        "assistant_capabilities": capability_map.get(assistant_mode, capability_map["basic"]),
        "assistant_sources": assistant_sources,
        "assistant_guardrails": assistant_guardrails,
    }


@bp.route("/assistant")
@login_required
def assistant():
    ctx = _get_plan_context(session.get("user_id"))
    return render_template("assistant.html",
                           assistant_access=ctx["assistant_access"],
                           assistant_mode=ctx["assistant_mode"],
                           assistant_mode_label=ctx["assistant_mode_label"],
                           plan_name=ctx["plan_name"],
                           assistant_capabilities=ctx["assistant_capabilities"],
                           assistant_sources=ctx["assistant_sources"],
                           assistant_guardrails=ctx["assistant_guardrails"])


@bp.route("/health", methods=["GET"])
def health():
    from .. import ai_service

    news_state = load_sync_state(current_app.config["NEWS_SYNC_STATE_FILE"])
    payload = {
        "status": "ok",
        "db": "up",
        "ai": "up" if ai_service.is_available else "degraded",
        "media_root": str(current_app.config["INCIDENT_MEDIA_ROOT"]),
        "news_sync": {
            "last_sync_at": news_state.get("last_sync_at"),
            "last_error": news_state.get("last_error"),
            "cooldown_minutes": current_app.config["NEWS_SYNC_MINUTES"],
        },
    }
    try:
        db.session.execute(text("SELECT 1"))
    except Exception as exc:
        payload["status"] = "degraded"
        payload["db"] = f"down: {exc}"
        return jsonify(payload), 503
    return jsonify(payload)


@bp.route("/api/admin/news-sources", methods=["GET"])
@login_required
def api_list_news_sources():
    from ..models import User

    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() != "admin":
        return jsonify({"error": "No autorizado"}), 403
    return jsonify(load_sources(current_app.config["NEWS_SOURCES_FILE"]))


@bp.route("/api/admin/news-sources", methods=["POST"])
@login_required
def api_save_news_sources():
    from ..models import User

    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() != "admin":
        return jsonify({"error": "No autorizado"}), 403

    payload = request.get_json(silent=True) or {}
    sources = payload.get("sources")
    if not isinstance(sources, list):
        return jsonify({"error": "sources debe ser una lista"}), 400
    return jsonify({"success": True, "sources": save_sources(current_app.config["NEWS_SOURCES_FILE"], sources)})


@bp.route("/api/admin/news-sync", methods=["POST"])
@login_required
def api_sync_news_sources():
    from ..models import User, Role, Incident, IncidentCategory

    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() != "admin":
        return jsonify({"error": "No autorizado"}), 403

    sources = load_sources(current_app.config["NEWS_SOURCES_FILE"])
    categories = {row.name: row.id for row in IncidentCategory.query.all()}
    try:
        cleanup_expired = cleanup_expired_news_incidents(db.session, Incident, current_app.config["NEWS_MANIFEST_FILE"])
        sync_result = sync_sources(
            db.session,
            (User, Role, Incident, IncidentCategory),
            categories,
            current_app.config["NEWS_MANIFEST_FILE"],
            sources,
        )
        db.session.commit()
        return jsonify({"success": True, "expired_removed": cleanup_expired, **sync_result})
    except Exception as exc:
        db.session.rollback()
        logger.warning("News sync failed: %s", exc)
        return jsonify({"error": f"No se pudo sincronizar fuentes: {exc}"}), 500


@bp.route("/api/admin/security-events", methods=["GET"])
@login_required
def api_security_events():
    from ..models import User

    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() != "admin":
        return jsonify({"error": "No autorizado"}), 403

    limit = request.args.get("limit", 100, type=int)
    return jsonify({
        "events": list_security_events(limit),
        "count": min(max(limit, 1), 500),
    })


@bp.route("/api/geo/search", methods=["POST"])
@login_required
def api_geo_search():
    data = request.get_json(silent=True) or {}
    query = (data.get("query") or "").strip()
    if not query:
        return jsonify({"error": "Lugar requerido"}), 400

    try:
        results = _geocode_place(query)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 503

    if not results:
        return jsonify({"error": "No encontré ese lugar"}), 404

    best = results[0]
    return jsonify({"query": query, "best_match": best, "results": results})


@bp.route("/api/geo/reverse", methods=["POST"])
@login_required
def api_geo_reverse():
    data = request.get_json(silent=True) or {}
    lat = data.get("lat")
    lng = data.get("lng")

    if lat is None or lng is None:
        return jsonify({"error": "Coordenadas requeridas"}), 400

    try:
        lat, lng = float(lat), float(lng)
    except (ValueError, TypeError):
        return jsonify({"error": "Coordenadas invalidas"}), 400

    place_name = _reverse_geocode(lat, lng)
    return jsonify({
        "lat": lat,
        "lng": lng,
        "place_name": place_name or f"{lat:.5f}, {lng:.5f}"
    })


# ============================================================================
# CATEGORÍAS (público)
# ============================================================================

@bp.route("/api/categories")
def categories():
    from ..models import IncidentCategory
    rows = IncidentCategory.query.all()
    return jsonify([{
        "id": c.id, "name": c.name, "icon": c.icon,
        "color": c.color_hex, "default_severity": c.default_severity
    } for c in rows])


# ============================================================================
# INCIDENTES — sesión Flask (no JWT)
# ============================================================================

def _serialize(row_dict):
    """Serializa un dict de fila SQL a JSON-safe."""
    d = dict(row_dict)
    for k, v in d.items():
        if hasattr(v, 'isoformat'):
            d[k] = v.isoformat()
        elif hasattr(v, '__float__') and not isinstance(v, (int, bool)):
            d[k] = float(v)
        elif v is None:
            d[k] = None
    return d


def _incident_public_id(payload):
    value = payload.get("incident_id") or payload.get("id")
    return int(value) if value is not None else None


def _viewer_is_authenticated():
    return bool(session.get("user_id"))


def _sanitize_public_comments(comments):
    safe_comments = []
    for comment in comments or []:
        safe_comments.append({
            "text": comment.get("text") or comment.get("body") or "",
            "username": "Usuario anonimo",
            "created_at": comment.get("created_at"),
        })
    return safe_comments


def _sanitize_public_incident(payload):
    safe = dict(payload)
    safe["reporter_username"] = None
    safe["username"] = None
    safe["reporter_name"] = None
    safe["reporter_id"] = None
    safe["user_id"] = None
    safe["reporter_profile"] = None
    if safe.get("comments"):
        safe["comments"] = _sanitize_public_comments(safe["comments"])
    return safe


def _digest_window_hours():
    return max(1, int(current_app.config.get("EMAIL_SUMMARY_HOURS", 24)))


def _digest_cutoff():
    return datetime.utcnow() - timedelta(hours=_digest_window_hours())


def _load_digest_incidents_for_user(user_id, limit=None):
    prefs = load_preferences(current_app.config["USER_PREFS_ROOT"], user_id)
    rows = db.session.execute(text("""
        SELECT
            i.id,
            i.title,
            i.description,
            i.severity,
            i.created_at,
            c.name AS category_name,
            u.username AS reporter_username
        FROM incidents i
        LEFT JOIN incident_categories c ON c.id = i.category_id
        LEFT JOIN users u ON u.id = i.user_id
        WHERE i.status <> 'dismissed'
          AND i.created_at >= :cutoff
        ORDER BY i.created_at DESC
        LIMIT :limit
    """), {
        "cutoff": _digest_cutoff(),
        "limit": int(limit or current_app.config.get("EMAIL_SUMMARY_MAX_ITEMS", 8)),
    }).mappings().all()

    allowed = set(prefs.get("incident_types") or [])
    incidents = []
    for row in rows:
        item = dict(row)
        if allowed and item.get("category_name") not in allowed:
            continue
        item["location_label"] = _extract_description_field(item.get("description"), "Zona objetivo") or "zona reportada"
        item["source_name"] = _incident_source_label(item)
        incidents.append(item)
    return incidents


def _build_summary_email_payload(user, incidents):
    from .. import ai_service

    digest = ai_service.summarize_incidents_digest(
        incidents,
        timeframe_hours=_digest_window_hours(),
        recipient_name=user.full_name or user.username,
    )
    headline = digest.get("headline") or "Resumen PanamaAlert"
    summary = digest.get("summary") or "Revisa las alertas recientes en PanamaAlert."
    bullets = digest.get("bullets") or []
    return {
        "subject": f"[{current_app.config.get('APP_NAME', 'PanamaAlert')}] {headline}",
        "text": (
            f"Hola {user.full_name or user.username},\n\n"
            f"{summary}\n\n"
            + ("\n".join(f"- {bullet}" for bullet in bullets) if bullets else "Sin detalles adicionales.")
            + "\n\nIngresa a PanamaAlert para ver el mapa completo y los pings recientes.\n"
        ),
        "html": f"""
        <html><body style="font-family:Segoe UI,Arial,sans-serif;padding:24px;background:#f8fafc;color:#0f172a;">
            <div style="max-width:720px;margin:0 auto;background:#fff;border:1px solid #e2e8f0;border-radius:18px;padding:28px;">
                <div style="font-size:12px;font-weight:800;letter-spacing:.04em;text-transform:uppercase;color:#0284c7;">PanamaAlert · Resumen IA</div>
                <h1 style="margin:12px 0 10px;font-size:28px;line-height:1.1;">{headline}</h1>
                <p style="color:#475569;line-height:1.75;">{summary}</p>
                <div style="margin-top:18px;padding:18px;border-radius:16px;background:#f8fafc;border:1px solid #e2e8f0;">
                    {''.join(f"<div style='padding:10px 0;border-bottom:1px solid #e2e8f0;color:#0f172a;line-height:1.6;'>{bullet}</div>" for bullet in bullets) or "<div style='color:#64748b;'>Sin alertas nuevas para resumir.</div>"}
                </div>
                <p style="margin-top:20px;color:#64748b;line-height:1.7;">Este correo se envía solo porque activaste las notificaciones por correo desde tu perfil.</p>
            </div>
        </body></html>
        """,
    }


def _load_incident_ai_map(incident_ids):
    if not incident_ids:
        return {}
    import json as json_mod

    placeholders = ",".join([str(int(i)) for i in incident_ids[:500]])
    ai_rows = db.session.execute(text(f"""
        SELECT incident_id, decision, confidence, reason, flags, alert_level
        FROM ai_analyses
        WHERE incident_id IN ({placeholders})
        AND id IN (
            SELECT MAX(id) FROM ai_analyses
            WHERE incident_id IN ({placeholders})
            GROUP BY incident_id
        )
    """)).mappings().all()
    ai_map = {}
    for a in ai_rows:
        flags_raw = a["flags"]
        flags_parsed = []
        if flags_raw:
            try:
                flags_parsed = json_mod.loads(flags_raw) if isinstance(flags_raw, str) else flags_raw
            except Exception:
                pass
        ai_map[int(a["incident_id"])] = {
            "decision": a["decision"],
            "confidence": float(a["confidence"]),
            "alert_level": a["alert_level"],
            "reason": a["reason"] or "",
            "flags": flags_parsed
        }
    return ai_map


def _load_incident_comments_map(incident_ids):
    if not incident_ids:
        return {}
    placeholders = ",".join([str(int(i)) for i in incident_ids[:500]])
    comment_rows = db.session.execute(text(f"""
        SELECT ic.incident_id, ic.body, ic.created_at, u.username
        FROM incident_comments ic
        JOIN users u ON ic.user_id = u.id
        WHERE ic.incident_id IN ({placeholders})
        ORDER BY ic.created_at ASC
    """)).mappings().all()

    comments_map = {}
    for c in comment_rows:
        iid = int(c["incident_id"])
        comments_map.setdefault(iid, []).append({
            "text": c["body"],
            "username": c["username"],
            "created_at": c["created_at"].isoformat() if hasattr(c["created_at"], "isoformat") else str(c["created_at"])
        })
    return comments_map


def _incident_base_query_sql():
    return """
        SELECT * FROM v_incidents_full
        WHERE expires_at IS NULL OR expires_at >= NOW()
    """


def _incident_base_fallback_sql():
    return """
        SELECT i.*, u.username AS reporter_username,
               ic.name AS category_name, ic.color_hex AS category_color
        FROM incidents i
        JOIN users u ON i.user_id = u.id
        JOIN incident_categories ic ON i.category_id = ic.id
        WHERE i.status != 'dismissed'
          AND (i.expires_at IS NULL OR i.expires_at >= NOW())
    """


def _incident_base_minimal_sql():
    return """
        SELECT i.*, u.username AS reporter_username,
               ic.name AS category_name, ic.color_hex AS category_color
        FROM incidents i
        JOIN users u ON i.user_id = u.id
        JOIN incident_categories ic ON i.category_id = ic.id
        WHERE i.status != 'dismissed'
    """


@bp.route("/api/incidents", methods=["GET"])
def api_incidents():
    """Lista TODOS los incidentes públicos (no dismissed). Incluye análisis IA."""
    try:
        from ..models import Incident, IncidentCategory, Role, User
        limit = min(max(request.args.get("limit", 250, type=int), 1), 500)
        categories = {row.name: row.id for row in IncidentCategory.query.all()}
        sync_meta = {"triggered": False, "reason": "disabled_during_listing"}

        try:
            cleanup_expired_news_incidents(db.session, Incident, current_app.config["NEWS_MANIFEST_FILE"])
            db.session.commit()
        except Exception as cleanup_exc:
            db.session.rollback()
            logger.warning("Expired news cleanup skipped while listing incidents: %s", cleanup_exc)
        # Intentar vista primero, fallback a tabla directa
        try:
            rows = db.session.execute(
                text(f"{_incident_base_query_sql()} ORDER BY created_at DESC LIMIT :limit"),
                {"limit": limit}
            ).mappings().all()
        except Exception:
            db.session.rollback()
            try:
                rows = db.session.execute(
                    text(f"{_incident_base_fallback_sql()} ORDER BY i.created_at DESC LIMIT :limit"),
                    {"limit": limit}
                ).mappings().all()
            except Exception:
                db.session.rollback()
                rows = db.session.execute(
                    text(f"{_incident_base_minimal_sql()} ORDER BY i.created_at DESC LIMIT :limit"),
                    {"limit": limit}
                ).mappings().all()

        results = [_serialize(r) for r in rows]

        incident_ids = [_incident_public_id(r) for r in results if _incident_public_id(r)]
        user_ids = [r.get("user_id") for r in results if r.get("user_id")]

        try:
            ai_map = _load_incident_ai_map(incident_ids)
            for r in results:
                rid = r.get("id") or r.get("incident_id")
                if rid and int(rid) in ai_map:
                    r["ai_analysis"] = ai_map[int(rid)]
        except Exception as ai_err:
            logger.debug(f"AI enrichment skipped: {ai_err}")

        try:
            reporter_profiles = build_reporter_profiles(db.session, user_ids) if user_ids else {}
            for r in results:
                uid = r.get("user_id")
                if uid in reporter_profiles:
                    r["reporter_profile"] = reporter_profiles[uid]
        except Exception as trust_err:
            logger.debug(f"Reporter profile enrichment skipped: {trust_err}")

        try:
            _build_incident_moderation_insights(results)
        except Exception as insight_err:
            logger.debug(f"Moderation insights skipped: {insight_err}")

        if not _viewer_is_authenticated():
            results = [_sanitize_public_incident(item) for item in results]

        return jsonify(results)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/incidents/<int:iid>", methods=["GET"])
def api_get_incident(iid):
    try:
        row = None
        try:
            row = db.session.execute(
                text(f"{_incident_base_query_sql()} AND incident_id = :id LIMIT 1"),
                {"id": iid}
            ).mappings().first()
        except Exception:
            db.session.rollback()
            try:
                row = db.session.execute(
                    text(f"{_incident_base_fallback_sql()} AND i.id = :id LIMIT 1"),
                    {"id": iid}
                ).mappings().first()
            except Exception:
                db.session.rollback()
                row = db.session.execute(
                    text(f"{_incident_base_minimal_sql()} AND i.id = :id LIMIT 1"),
                    {"id": iid}
                ).mappings().first()

        if not row:
            return jsonify({"error": "Incidente no encontrado"}), 404

        result = _serialize(row)
        incident_ids = [_incident_public_id(result)]
        user_ids = [result.get("user_id")] if result.get("user_id") else []

        try:
            ai_map = _load_incident_ai_map(incident_ids)
            if iid in ai_map:
                result["ai_analysis"] = ai_map[iid]
        except Exception as ai_err:
            logger.debug(f"Incident detail AI enrichment skipped: {ai_err}")

        try:
            reporter_profiles = build_reporter_profiles(db.session, user_ids) if user_ids else {}
            uid = result.get("user_id")
            if uid in reporter_profiles:
                result["reporter_profile"] = reporter_profiles[uid]
        except Exception as trust_err:
            logger.debug(f"Incident detail reporter profile skipped: {trust_err}")

        try:
            comments_map = _load_incident_comments_map(incident_ids)
            result["comments"] = comments_map.get(iid, [])
        except Exception as cmt_err:
            logger.debug(f"Incident detail comments skipped: {cmt_err}")

        try:
            result["evidence"] = list_incident_evidence(current_app.config["INCIDENT_MEDIA_ROOT"], iid)
            result["analysis_explainability"] = explain_analysis(
                result,
                result.get("ai_analysis"),
                result.get("reporter_profile"),
            )
        except Exception as detail_err:
            logger.debug(f"Incident detail evidence/explainability skipped: {detail_err}")

        try:
            _build_incident_moderation_insights([result])
        except Exception as insight_err:
            logger.debug(f"Incident detail moderation insights skipped: {insight_err}")

        if not _viewer_is_authenticated():
            result = _sanitize_public_incident(result)

        return jsonify(result)
    except Exception as exc:
        logger.exception("Incident detail failed for %s", iid)
        return jsonify({"error": str(exc)}), 500


@bp.route("/api/incidents", methods=["POST"])
@login_required
def api_create_incident():
    """Crea incidente vía SP. Respeta límites del plan del usuario."""
    b = request.get_json(silent=True) or {}
    try:
        lat, lng = float(b["lat"]), float(b["lng"])
    except (KeyError, TypeError, ValueError):
        return jsonify({"error": "lat/lng requeridos"}), 400
    try:
        _validate_panama_coords(lat, lng)
    except ValueError:
        return jsonify({"error": "Ubicacion fuera del territorio admitido de Panama"}), 400

    try:
        title = _clean_text(b.get("title"), "Titulo", 160, min_len=5)
        desc = _clean_text(b.get("description"), "Descripcion", 1000, min_len=15)
        severity = max(1, min(5, int(b.get("severity") or 3)))
        category_id = int(b.get("category_id") or 1)
    except (TypeError, ValueError):
        return jsonify({"error": "Datos del reporte invalidos"}), 400

    # --- Enforce plan limits ---
    try:
        from ..models import Subscription, Plan
        sub = Subscription.query.filter_by(user_id=session["user_id"]).first()
        plan = Plan.query.get(sub.plan_id) if sub else None
        max_per_day = plan.max_alerts_per_day if plan else 10
        today_count = db.session.execute(text(
            "SELECT COUNT(*) FROM incidents WHERE user_id = :uid AND DATE(created_at) = CURDATE()"
        ), {"uid": session["user_id"]}).scalar() or 0
        if today_count >= max_per_day:
            return jsonify({
                "error": f"Has alcanzado el limite de {max_per_day} reportes por dia de tu plan {plan.name if plan else 'Free'}. Mejora tu plan para reportar mas."
            }), 429
    except Exception as limit_err:
        logger.debug(f"Plan limit check skipped: {limit_err}")

    try:
        conn = db.session.connection()
        conn.exec_driver_sql("SET @new_id = 0")
        conn.exec_driver_sql(
            "CALL sp_create_incident(%s,%s,%s,%s,%s,%s,%s,%s,@new_id)",
            (session["user_id"], category_id, None, title, desc, lat, lng, severity)
        )
        new_id = conn.exec_driver_sql("SELECT @new_id").scalar()
        db.session.commit()

        incident_id = int(new_id)

        # --- AI Smart Moderation (FakePingDetector + Content Moderation pipeline) ---
        ai_result = None
        try:
            from .. import ai_service
            svc = ai_service
            from ..models import IncidentCategory
            cat = IncidentCategory.query.get(category_id)
            cat_name = cat.name if cat else "General"

            if hasattr(current_app, 'smart_moderator') and current_app.smart_moderator:
                # Use SmartModerator pipeline (combines fake detection + content moderation)
                incident_data = {
                    "title": title, "description": desc, "category": cat_name,
                    "lat": lat, "lng": lng, "severity": severity
                }

                # Get user history for fake detection
                user_history = db.session.execute(text("""
                    SELECT id, title, lat, lng, severity, created_at
                    FROM incidents WHERE user_id = :uid
                    ORDER BY created_at DESC LIMIT 10
                """), {"uid": session["user_id"]}).mappings().all()
                user_hist = [{
                    "id": r["id"], "title": r["title"], "lat": float(r["lat"]),
                    "lng": float(r["lng"]), "severity": int(r["severity"]),
                    "created_at": str(r["created_at"])
                } for r in user_history]

                sm_result = current_app.smart_moderator.moderate(incident_data, user_hist, [])
                ai_result = {
                    "decision": sm_result.get("decision", "review"),
                    "confidence": sm_result.get("confidence", 0.5),
                    "reason": sm_result.get("reason", ""),
                    "flags": sm_result.get("flags", []),
                    "alert_level": sm_result.get("alert_level", "none"),
                    "risk_score": sm_result.get("fake_analysis", {}).get("risk_score", 0),
                    "is_fake": sm_result.get("fake_analysis", {}).get("is_fake", False),
                    "signals": sm_result.get("fake_analysis", {}).get("top_signals", [])
                }
                svc.save_analysis(incident_id, ai_result)
                logger.info(f"SmartModerator for incident {incident_id}: {ai_result['decision']} (risk={sm_result.get('risk_score', 0)})")
            elif svc.is_available:
                # Fallback to basic AI moderation
                ai_result = svc.moderate_incident(title, desc, cat_name, lat, lng)
                svc.save_analysis(incident_id, ai_result)
                logger.info(f"AI moderation for incident {incident_id}: {ai_result.get('decision')}")
        except Exception as ai_err:
            logger.warning(f"AI moderation skipped for incident {incident_id}: {ai_err}")

        # --- Geofence alerts (with AI-generated message if available) ---
        try:
            _check_geofence_alerts(incident_id, {
                "title": title, "lat": lat, "lng": lng,
                "description": desc, "category": cat_name if 'cat_name' in dir() else "General",
                "severity": severity,
                "alert_level": ai_result.get("alert_level", "medium") if ai_result else "medium"
            })
        except Exception:
            pass

        response = {"success": True, "id": incident_id}
        if ai_result:
            response["ai_analysis"] = {
                "decision": ai_result.get("decision"),
                "confidence": ai_result.get("confidence"),
                "alert_level": ai_result.get("alert_level")
            }
        return jsonify(response), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/incidents/<int:iid>", methods=["PUT"])
@login_required
def api_update_incident(iid):
    b = request.get_json(silent=True) or {}
    sets, params = [], {"id": iid, "uid": session["user_id"]}
    if "title" in b:
        try:
            params["t"] = _clean_text(b["title"], "Titulo", 160, min_len=5)
            sets.append("title=:t")
        except ValueError:
            return jsonify({"error": "Titulo invalido"}), 400
    if "description" in b:
        try:
            params["d"] = _clean_text(b["description"], "Descripcion", 1000, min_len=15)
            sets.append("description=:d")
        except ValueError:
            return jsonify({"error": "Descripcion invalida"}), 400
    if "severity" in b:
        try:
            params["s"] = max(1, min(5, int(b["severity"])))
            sets.append("severity=:s")
        except (TypeError, ValueError):
            return jsonify({"error": "Severidad invalida"}), 400
    if not sets:
        return jsonify({"error": "Nada que actualizar"}), 400
    try:
        db.session.execute(
            text(f"UPDATE incidents SET {','.join(sets)} WHERE id=:id AND user_id=:uid"),
            params
        )
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/incidents/<int:iid>", methods=["DELETE"])
@login_required
def api_delete_incident(iid):
    try:
        db.session.execute(
            text("DELETE FROM incidents WHERE id=:id AND user_id=:uid"),
            {"id": iid, "uid": session["user_id"]}
        )
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/incidents/<int:iid>/vote", methods=["POST"])
@login_required
def api_vote(iid):
    vote = (request.get_json(silent=True) or {}).get("vote")
    if vote not in (1, -1):
        return jsonify({"error": "vote debe ser 1 o -1"}), 400
    try:
        db.session.execute(
            text("""INSERT INTO incident_votes(incident_id, user_id, vote)
                    VALUES(:inc, :uid, :v)
                    ON DUPLICATE KEY UPDATE vote=:v"""),
            {"inc": iid, "uid": session["user_id"], "v": vote}
        )
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


# ============================================================================
# COMENTARIOS — columna real: body (NO text)
# ============================================================================

@bp.route("/api/incidents/<int:iid>/comments", methods=["POST"])
@login_required
def api_create_comment(iid):
    b = request.get_json(silent=True) or {}
    body_val = (b.get("text") or "").strip()
    if not body_val or len(body_val) > 1000:
        return jsonify({"error": "Comentario inválido (1-1000 chars)"}), 400
    try:
        from ..models import IncidentComment
        cmt = IncidentComment(incident_id=iid, user_id=session["user_id"], body=body_val)
        db.session.add(cmt)
        db.session.commit()
        return jsonify({"success": True, "id": cmt.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/incidents/<int:iid>/media", methods=["POST"])
@login_required
def api_upload_incident_media(iid):
    from ..models import Incident, User

    incident = Incident.query.get(iid)
    if not incident:
        return jsonify({"error": "Incidente no encontrado"}), 404

    user = User.query.get(session.get("user_id"))
    is_privileged = user and user.role_name() in ("admin", "moderator")
    if incident.user_id != session.get("user_id") and not is_privileged:
        return jsonify({"error": "No autorizado"}), 403

    file_obj = request.files.get("file")
    if not file_obj:
        return jsonify({"error": "Archivo requerido"}), 400

    try:
        item = save_incident_evidence(
            file_obj,
            iid,
            session.get("user_id"),
            current_app.config["INCIDENT_MEDIA_ROOT"],
            current_app.config["INCIDENT_MEDIA_MAX_BYTES"],
            current_app.config["INCIDENT_MEDIA_ALLOWED_EXTENSIONS"],
            current_app.config["INCIDENT_MEDIA_MAX_FILES"],
        )
        return jsonify({"success": True, "item": item}), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        logger.warning("Incident media upload failed: %s", exc)
        return jsonify({"error": "No se pudo guardar la evidencia"}), 500


# ============================================================================
# MODERACIÓN — status válidos: pending, verified, resolved, dismissed
# ============================================================================

@bp.route("/api/incidents/<int:iid>/status", methods=["PUT"])
@login_required
def api_update_status(iid):
    from ..models import User
    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() not in ("admin", "moderator"):
        return jsonify({"error": "No autorizado"}), 403

    b = request.get_json(silent=True) or {}
    status = b.get("status")
    if status not in ("pending", "verified", "resolved", "dismissed"):
        return jsonify({"error": "Status inválido"}), 400

    try:
        if status == "verified":
            db.session.execute(
                text("UPDATE incidents SET status=:s, verified_by=:vb, verified_at=NOW() WHERE id=:id"),
                {"s": status, "vb": session["user_id"], "id": iid}
            )
        else:
            db.session.execute(
                text("UPDATE incidents SET status=:s WHERE id=:id"),
                {"s": status, "id": iid}
            )
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/incidents/<int:iid>/admin-review", methods=["PUT"])
@login_required
def api_admin_review_incident(iid):
    from ..models import Incident, User

    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() not in ("admin", "moderator"):
        return jsonify({"error": "No autorizado"}), 403

    incident = Incident.query.get(iid)
    if not incident:
        return jsonify({"error": "Incidente no encontrado"}), 404

    payload = request.get_json(silent=True) or {}
    updates = []
    params = {"id": iid}

    if "lat" in payload or "lng" in payload:
        try:
            lat = float(payload.get("lat"))
            lng = float(payload.get("lng"))
            _validate_panama_coords(lat, lng)
            params["lat"] = lat
            params["lng"] = lng
            updates.extend(["lat=:lat", "lng=:lng"])
        except (TypeError, ValueError) as exc:
            return jsonify({"error": str(exc) or "Coordenadas invalidas"}), 400

    if "title" in payload:
        try:
            params["title"] = _clean_text(payload.get("title"), "Titulo", 160, min_len=5)
            updates.append("title=:title")
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    if "description" in payload:
        try:
            params["description"] = _clean_text(payload.get("description"), "Descripcion", 1000, min_len=15)
            updates.append("description=:description")
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    if "status" in payload:
        status = payload.get("status")
        if status not in ("pending", "verified", "resolved", "dismissed"):
            return jsonify({"error": "Status invalido"}), 400
        params["status"] = status
        updates.append("status=:status")
        if status == "verified":
            params["verified_by"] = session.get("user_id")
            updates.extend(["verified_by=:verified_by", "verified_at=NOW()"])

    if not updates:
        return jsonify({"error": "Nada que actualizar"}), 400

    try:
        db.session.execute(text(f"UPDATE incidents SET {', '.join(updates)} WHERE id=:id"), params)
        db.session.commit()
        return jsonify({"success": True, "updated": True})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500


@bp.route("/api/incidents/<int:iid>/mark-duplicate", methods=["POST"])
@login_required
def api_mark_duplicate(iid):
    from ..models import Incident, User

    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() not in ("admin", "moderator"):
        return jsonify({"error": "No autorizado"}), 403

    payload = request.get_json(silent=True) or {}
    duplicate_of = payload.get("duplicate_of")
    note = (payload.get("note") or "").strip()
    if not duplicate_of:
        return jsonify({"error": "Debes indicar el incidente principal"}), 400

    duplicate = Incident.query.get(iid)
    primary = Incident.query.get(int(duplicate_of))
    if not duplicate or not primary:
        return jsonify({"error": "Incidente no encontrado"}), 404
    if duplicate.id == primary.id:
        return jsonify({"error": "No puedes marcar un incidente como duplicado de si mismo"}), 400

    extra_note = f"[Moderacion: marcado como duplicado de #{primary.id}"
    if note:
        extra_note += f" | {note[:180]}"
    extra_note += "]"
    updated_description = (duplicate.description or "").strip()
    if extra_note not in updated_description:
        updated_description = f"{updated_description}\n{extra_note}".strip()

    try:
        db.session.execute(
            text("UPDATE incidents SET status='dismissed', description=:description WHERE id=:id"),
            {"id": iid, "description": updated_description}
        )
        db.session.commit()
        return jsonify({"success": True, "duplicate_of": int(primary.id)})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500


# ============================================================================
# USER PROFILE — campos reales del modelo
# ============================================================================

@bp.route("/api/user/profile", methods=["GET"])
@login_required
def api_user_profile():
    from ..models import User, Subscription, Plan, AlertSubscription, IncidentCategory
    user = User.query.get(session.get("user_id"))
    if not user:
        return jsonify({"error": "Usuario no encontrado"}), 404

    plan_ctx = _get_plan_context(user.id)
    sub = plan_ctx['subscription']
    plan = plan_ctx['plan']

    # Contar geofences activos
    geo_used = AlertSubscription.query.filter_by(
        user_id=user.id, active=True
    ).count()

    # Daily usage count
    today_count = db.session.execute(text(
        "SELECT COUNT(*) FROM incidents WHERE user_id = :uid AND DATE(created_at) = CURDATE()"
    ), {"uid": user.id}).scalar() or 0
    reporter_profile = build_reporter_profiles(db.session, [user.id]).get(user.id, {})
    notification_prefs = load_preferences(current_app.config["USER_PREFS_ROOT"], user.id)
    categories = IncidentCategory.query.order_by(IncidentCategory.name.asc()).all()

    return jsonify({
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "role": user.role_name(),
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "plan_name": plan.name if plan else "Free",
        "plan_price": float(plan.price_monthly_usd) if plan and plan.price_monthly_usd else 0,
        "plan_id": plan.id if plan else None,
        "max_alerts_per_day": plan.max_alerts_per_day if plan else 10,
        "alerts_today": today_count,
        "geofences_limit": plan.max_geo_fences if plan else 1,
        "geofences_used": geo_used,
        "api_access": plan.api_access if plan else False,
        "priority_support": plan.priority_support if plan else False,
        "features": plan.features_json if plan else None,
        "assistant_access": plan_ctx["assistant_access"],
        "assistant_mode": plan_ctx["assistant_mode"],
        "assistant_mode_label": plan_ctx["assistant_mode_label"],
        "subscription_status": sub.status if sub else "none",
        "subscription_expiry": sub.expires_at.isoformat() if sub and sub.expires_at else None,
        "reporter_profile": reporter_profile,
        "notification_preferences": notification_prefs,
        "notification_categories": [c.name for c in categories]
    })


@bp.route("/api/user/preferences", methods=["GET"])
@login_required
def api_get_preferences():
    prefs = load_preferences(current_app.config["USER_PREFS_ROOT"], session.get("user_id"))
    return jsonify(prefs)


@bp.route("/api/user/preferences", methods=["POST"])
@login_required
def api_save_preferences():
    payload = request.get_json(silent=True) or {}
    prefs = save_preferences(current_app.config["USER_PREFS_ROOT"], session.get("user_id"), payload)
    return jsonify({"success": True, "preferences": prefs})


@bp.route("/api/user/preferences/test-email", methods=["POST"])
@login_required
def api_send_test_email():
    from ..models import User
    try:
        user = User.query.get(session.get("user_id"))
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404

        prefs = load_preferences(current_app.config["USER_PREFS_ROOT"], user.id)
        if not prefs.get("email_enabled"):
            return jsonify({"error": "Activa primero las notificaciones por correo en tu perfil"}), 400
        if not mail_is_configured():
            return jsonify({"error": "El servidor SMTP no esta configurado"}), 503

        sent = send_email(
            to_email=user.email,
            subject=f"[{current_app.config.get('APP_NAME', 'PanamaAlert')}] Prueba de correo",
            text_body=(
                f"Hola {user.full_name or user.username},\n\n"
                "Este es un correo de prueba de PanamaAlert. Si lo recibiste, "
                "el envio SMTP ya esta funcionando correctamente.\n"
            ),
            html_body=f"""
            <html><body style="font-family:Segoe UI,Arial,sans-serif;padding:24px;background:#f8fafc;color:#0f172a;">
            <div style="max-width:620px;margin:0 auto;background:#fff;border:1px solid #e2e8f0;border-radius:16px;padding:28px;">
                <div style="font-size:12px;font-weight:800;letter-spacing:.04em;text-transform:uppercase;color:#0284c7;">PanamaAlert</div>
                <h1 style="margin:12px 0 8px;font-size:28px;">Correo de prueba enviado correctamente</h1>
                <p style="color:#475569;line-height:1.7;">Hola {user.full_name or user.username}, este es un correo de prueba para confirmar que las notificaciones por email ya funcionan de verdad.</p>
            </div>
            </body></html>
            """,
        )
        if not sent:
            return jsonify({"error": "No se pudo enviar el correo de prueba"}), 500
        return jsonify({"success": True, "message": f"Correo enviado a {user.email}"})
    except Exception as exc:
        logger.exception("Test email endpoint failed: %s", exc)
        return jsonify({"error": "El servidor devolvió un error interno al enviar el correo de prueba"}), 500


@bp.route("/api/user/preferences/send-summary", methods=["POST"])
@login_required
def api_send_summary_email():
    from ..models import User
    try:
        user = User.query.get(session.get("user_id"))
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404

        prefs = load_preferences(current_app.config["USER_PREFS_ROOT"], user.id)
        if not prefs.get("email_enabled"):
            return jsonify({"error": "Activa primero las notificaciones por correo en tu perfil"}), 400
        if not mail_is_configured():
            return jsonify({"error": "El servidor SMTP no esta configurado"}), 503

        incidents = _load_digest_incidents_for_user(user.id)
        email_payload = _build_summary_email_payload(user, incidents)
        sent = send_email(
            to_email=user.email,
            subject=email_payload["subject"],
            text_body=email_payload["text"],
            html_body=email_payload["html"],
        )
        if not sent:
            return jsonify({"error": "No se pudo enviar el resumen al correo"}), 500

        prefs["last_summary_sent_at"] = datetime.utcnow().isoformat()
        save_preferences(current_app.config["USER_PREFS_ROOT"], user.id, prefs)
        return jsonify({
            "success": True,
            "message": f"Resumen enviado a {user.email}",
            "count": len(incidents),
        })
    except Exception as exc:
        logger.exception("Summary email endpoint failed: %s", exc)
        return jsonify({"error": "El servidor devolvió un error interno al generar el resumen"}), 500


@bp.route("/api/user/api-keys", methods=["GET"])
@login_required
def api_list_keys():
    from ..models import ApiKey
    keys = ApiKey.query.filter_by(user_id=session.get("user_id")).filter(
        ApiKey.revoked_at.is_(None)
    ).all()
    return jsonify([{
        "id": k.id,
        "name": k.name,
        "key_hash": k.key_hash[:16],
        "created_at": k.created_at.isoformat() if k.created_at else None
    } for k in keys])


@bp.route("/api/user/api-keys", methods=["POST"])
@login_required
def api_create_key():
    from ..models import ApiKey
    from ..security import hash_api_key
    import secrets

    plan_ctx = _get_plan_context(session.get("user_id"))
    if not plan_ctx['plan'] or not plan_ctx['plan'].api_access:
        return jsonify({"error": "Tu plan actual no incluye acceso API"}), 403

    key = secrets.token_urlsafe(32)
    api_key = ApiKey(
        user_id=session.get("user_id"),
        key_hash=hash_api_key(key),
        name=f"key_{secrets.token_hex(4)}"
    )
    try:
        db.session.add(api_key)
        db.session.commit()
        return jsonify({"key": key, "id": api_key.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/user/api-keys/<int:kid>", methods=["DELETE"])
@login_required
def api_delete_key(kid):
    try:
        db.session.execute(
            text("UPDATE api_keys SET revoked_at=NOW() WHERE id=:id AND user_id=:uid"),
            {"id": kid, "uid": session.get("user_id")}
        )
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/user/change-password", methods=["POST"])
@login_required
def api_change_password():
    from ..models import User
    from ..security import hash_password, valid_password

    b = request.get_json(silent=True) or {}
    password = b.get("password", "").strip()

    if not valid_password(password):
        return jsonify({"error": "8+ chars, mayúscula, minúscula y número"}), 400

    try:
        user = User.query.get(session.get("user_id"))
        user.password_hash = hash_password(password)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


def _luhn_check(card_number):
    """Algoritmo de Luhn para validar números de tarjeta."""
    digits = [int(d) for d in card_number if d.isdigit()]
    if len(digits) < 13 or len(digits) > 19:
        return False
    checksum = 0
    for i, d in enumerate(reversed(digits)):
        if i % 2 == 1:
            d *= 2
            if d > 9:
                d -= 9
        checksum += d
    return checksum % 10 == 0


def _detect_card_brand(number):
    """Detecta la marca de tarjeta basado en el BIN."""
    n = number.replace(" ", "").replace("-", "")
    if n.startswith("4"):
        return "Visa"
    if n[:2] in ("51", "52", "53", "54", "55"):
        return "Mastercard"
    if n[:2] in ("34", "37"):
        return "Amex"
    if n[:4] in ("6011", "6221", "6229"):
        return "Discover"
    return "Otra"


@bp.route("/api/user/payment-method", methods=["POST"])
@login_required
def api_save_payment_method():
    """Registra método de pago simulado con validación profesional (sin cobro real)."""
    import re
    b = request.get_json(silent=True) or {}

    card_number = re.sub(r"\s|-", "", b.get("card_number", "").strip())
    card_name = b.get("card_name", "").strip()
    card_expiry = b.get("card_expiry", "").strip()
    card_cvv = b.get("card_cvv", "").strip()

    errors = {}
    if not card_number or not card_number.isdigit():
        errors["card_number"] = "Número de tarjeta inválido"
    elif not _luhn_check(card_number):
        errors["card_number"] = "Número de tarjeta no pasa validación Luhn"

    if not card_name or len(card_name) < 3:
        errors["card_name"] = "Nombre del titular requerido (mínimo 3 caracteres)"

    if not re.match(r"^(0[1-9]|1[0-2])/\d{2}$", card_expiry):
        errors["card_expiry"] = "Formato de expiración inválido (MM/YY)"
    else:
        month, year = int(card_expiry[:2]), int("20" + card_expiry[3:])
        from datetime import date
        today = date.today()
        if year < today.year or (year == today.year and month < today.month):
            errors["card_expiry"] = "La tarjeta está expirada"

    if not re.match(r"^\d{3,4}$", card_cvv):
        errors["card_cvv"] = "CVV inválido (3 o 4 dígitos)"

    if errors:
        return jsonify({"error": "Datos inválidos", "fields": errors}), 400

    brand = _detect_card_brand(card_number)
    last4 = card_number[-4:]

    try:
        # Auto-fix: ensure payment_methods table has all required columns
        try:
            db.session.execute(text("""
                SELECT card_brand FROM payment_methods LIMIT 0
            """))
        except Exception:
            db.session.rollback()
            for col_sql in [
                "ALTER TABLE payment_methods ADD COLUMN card_brand VARCHAR(20) DEFAULT 'unknown' AFTER card_last4",
                "ALTER TABLE payment_methods ADD COLUMN card_name VARCHAR(60) NOT NULL DEFAULT 'Titular' AFTER card_brand",
                "ALTER TABLE payment_methods ADD COLUMN card_expiry VARCHAR(5) NOT NULL DEFAULT '00/00' AFTER card_name",
                "ALTER TABLE payment_methods ADD COLUMN updated_at DATETIME AFTER created_at",
            ]:
                try:
                    db.session.execute(text(col_sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        from ..models import PaymentMethod
        existing = PaymentMethod.query.filter_by(user_id=session["user_id"]).first()
        if existing:
            existing.card_last4 = last4
            existing.card_brand = brand
            existing.card_name = card_name[:60]
            existing.card_expiry = card_expiry
            existing.updated_at = __import__("datetime").datetime.utcnow()
        else:
            pm = PaymentMethod(
                user_id=session["user_id"],
                card_last4=last4,
                card_brand=brand,
                card_name=card_name[:60],
                card_expiry=card_expiry,
            )
            db.session.add(pm)
        db.session.commit()
        return jsonify({"success": True, "card_last4": last4, "card_brand": brand})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/user/payment-method", methods=["GET"])
@login_required
def api_get_payment_method():
    """Devuelve el método de pago registrado (enmascarado)."""
    from ..models import PaymentMethod
    pm = PaymentMethod.query.filter_by(user_id=session["user_id"]).first()
    if not pm:
        return jsonify({"has_card": False})
    return jsonify({
        "has_card": True,
        "card_last4": pm.card_last4,
        "card_brand": pm.card_brand,
        "card_name": pm.card_name,
        "card_expiry": pm.card_expiry,
        "created_at": pm.created_at.isoformat() if pm.created_at else None,
    })


@bp.route("/api/user/payment-method", methods=["DELETE"])
@login_required
def api_delete_payment_method():
    """Elimina el método de pago registrado."""
    from ..models import PaymentMethod
    pm = PaymentMethod.query.filter_by(user_id=session["user_id"]).first()
    if pm:
        db.session.delete(pm)
        db.session.commit()
    return jsonify({"success": True})


# ============================================================================
# GEOFENCES — columnas reales: center_lat, center_lng, active (NO latitude/longitude/deleted_at)
# ============================================================================

@bp.route("/api/alert-subscriptions", methods=["POST"])
@login_required
def api_create_alert_sub():
    from ..models import AlertSubscription
    b = request.get_json(silent=True) or {}

    plan_ctx = _get_plan_context(session.get("user_id"))
    plan = plan_ctx['plan']
    try:
        lat = float(b.get("center_lat") or b.get("latitude"))
        lng = float(b.get("center_lng") or b.get("longitude"))
        radius = _normalize_radius(b.get("radius_km", 5), min_value=0.5, max_value=25)
    except (KeyError, TypeError, ValueError):
        return jsonify({"error": "center_lat, center_lng requeridos"}), 400
    try:
        _validate_panama_coords(lat, lng)
    except ValueError:
        return jsonify({"error": "La zona debe estar dentro de Panama"}), 400

    current_count = AlertSubscription.query.filter_by(user_id=session["user_id"], active=True).count()
    geo_limit = plan.max_geo_fences if plan else 1
    if current_count >= geo_limit:
        return jsonify({"error": f"Tu plan permite un máximo de {geo_limit} zona(s) de alerta"}), 403

    try:
        sub = AlertSubscription(
            user_id=session["user_id"],
            center_lat=lat,
            center_lng=lng,
            radius_km=radius,
            min_severity=max(1, min(5, int(b.get("min_severity", 1)))),
            active=True
        )
        db.session.add(sub)
        db.session.commit()
        return jsonify({"success": True, "id": sub.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/alert-subscriptions", methods=["GET"])
@login_required
def api_list_alert_subs():
    from ..models import AlertSubscription
    subs = AlertSubscription.query.filter_by(
        user_id=session["user_id"], active=True
    ).order_by(AlertSubscription.created_at.desc()).all()

    return jsonify([{
        "id": s.id,
        "center_lat": float(s.center_lat),
        "center_lng": float(s.center_lng),
        "radius_km": float(s.radius_km),
        "min_severity": s.min_severity,
        "created_at": s.created_at.isoformat() if s.created_at else None
    } for s in subs])


@bp.route("/api/alert-subscriptions/<int:sub_id>", methods=["DELETE"])
@login_required
def api_delete_alert_sub(sub_id):
    from ..models import AlertSubscription
    try:
        sub = AlertSubscription.query.filter_by(
            id=sub_id, user_id=session["user_id"]
        ).first()
        if sub:
            sub.active = False
            db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


# ============================================================================
# NOTIFICACIONES — campos reales: type, message, is_read (NO title/data/read_at)
# ============================================================================

@bp.route("/api/notifications", methods=["GET"])
@login_required
def api_notifications():
    from ..models import Notification
    try:
        notifs = Notification.query.filter_by(
            user_id=session["user_id"], is_read=False
        ).order_by(Notification.created_at.desc()).limit(20).all()

        return jsonify([{
            "id": n.id,
            "type": n.type,
            "message": n.message,
            "incident_id": n.incident_id,
            "created_at": n.created_at.isoformat() if n.created_at else None
        } for n in notifs])
    except Exception:
        return jsonify([])


@bp.route("/api/notifications/<int:notif_id>/read", methods=["POST"])
@login_required
def api_mark_notif_read(notif_id):
    from ..models import Notification
    try:
        notif = Notification.query.filter_by(
            id=notif_id, user_id=session["user_id"]
        ).first()
        if notif:
            notif.is_read = True
            db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


# ============================================================================
# DASHBOARD — status correcto: dismissed (NO rejected)
# ============================================================================

@bp.route("/api/dashboard/stats", methods=["GET"])
@login_required
def api_dashboard_stats():
    try:
        total = db.session.execute(text("SELECT COUNT(*) FROM incidents")).scalar()
        verified = db.session.execute(
            text("SELECT COUNT(*) FROM incidents WHERE status='verified'")).scalar()
        pending = db.session.execute(
            text("SELECT COUNT(*) FROM incidents WHERE status='pending'")).scalar()
        dismissed = db.session.execute(
            text("SELECT COUNT(*) FROM incidents WHERE status='dismissed'")).scalar()

        cats = db.session.execute(text("""
            SELECT ic.name, COUNT(i.id) as count, ic.color_hex as color
            FROM incidents i
            JOIN incident_categories ic ON i.category_id = ic.id
            GROUP BY i.category_id, ic.name, ic.color_hex
            ORDER BY count DESC LIMIT 10
        """)).mappings().all()

        sevs = db.session.execute(text("""
            SELECT severity, COUNT(*) as count
            FROM incidents GROUP BY severity ORDER BY severity
        """)).mappings().all()

        severity = [0] * 5
        for s in sevs:
            idx = int(s['severity']) - 1
            if 0 <= idx < 5:
                severity[idx] = int(s['count'])

        hotspots = db.session.execute(text("""
            SELECT
                ROUND(i.lat, 2) as lat, ROUND(i.lng, 2) as lng,
                COUNT(*) as count,
                SUM(CASE WHEN i.created_at >= DATE_SUB(NOW(), INTERVAL 24 HOUR)
                    THEN 1 ELSE 0 END) as count_24h,
                ROUND(AVG(i.severity), 1) as avg_severity
            FROM incidents i
            GROUP BY ROUND(i.lat, 2), ROUND(i.lng, 2)
            ORDER BY count DESC LIMIT 10
        """)).mappings().all()

        return jsonify({
            "total": total or 0,
            "verified": verified or 0,
            "pending": pending or 0,
            "dismissed": dismissed or 0,
            "categories": [dict(c) for c in cats],
            "severity": severity,
            "hotspots": [{
                "lat": float(h['lat']), "lng": float(h['lng']),
                "count": int(h['count']), "count_24h": int(h['count_24h'] or 0),
                "avg_severity": float(h['avg_severity'] or 3),
                "location": f"({float(h['lat']):.2f}, {float(h['lng']):.2f})",
                "coords": f"{float(h['lat']):.4f}, {float(h['lng']):.4f}"
            } for h in hotspots]
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ============================================================================
# GEOFENCE ALERTAS — campos reales del modelo Notification
# ============================================================================

def _check_geofence_alerts(incident_id, incident_data):
    """Crea notificaciones para usuarios con geofences afectados. Usa IA si disponible."""
    try:
        lat = float(incident_data.get("lat", 0))
        lng = float(incident_data.get("lng", 0))

        rows = db.session.execute(text("""
            SELECT DISTINCT asub.user_id, asub.id as sub_id,
                   u.email,
                   u.full_name,
                   ST_Distance_Sphere(
                        POINT(asub.center_lng, asub.center_lat),
                        POINT(:lng, :lat)
                    ) / 1000 AS distance_km
            FROM alert_subscriptions asub
            JOIN users u ON u.id = asub.user_id
            WHERE asub.active = 1
            AND ST_Distance_Sphere(
                POINT(asub.center_lng, asub.center_lat),
                POINT(:lng, :lat)
            ) / 1000 <= asub.radius_km
        """), {"lat": lat, "lng": lng}).mappings().all()

        if not rows:
            return

        title = incident_data.get("title", "Nuevo incidente")

        # Try AI-generated alert message
        alert_message = None
        try:
            from .. import ai_service
            svc = ai_service
            if svc.is_available:
                ai_msg = svc.generate_alert_message(incident_data)
                if ai_msg and ai_msg.get("message"):
                    alert_message = f"{ai_msg.get('emoji', '⚠️')} {ai_msg['message']}"
        except Exception:
            pass

        for row in rows:
            from ..models import Notification
            prefs = load_preferences(current_app.config["USER_PREFS_ROOT"], row["user_id"])
            if not should_notify(prefs, incident_data):
                continue
            distance = round(float(row.get('distance_km', 0)), 1)
            msg = alert_message or f"Nuevo incidente: {title} - cerca de tu zona de alerta"
            if distance:
                msg += f" ({distance} km)"

            if prefs.get("push_enabled", True):
                notif = Notification(
                    user_id=row['user_id'],
                    incident_id=incident_id,
                    type="geofence",
                    message=msg[:500]
                )
                db.session.add(notif)
            if prefs.get("email_enabled") and row.get("email"):
                send_geofence_alert_email(
                    to_email=row["email"],
                    full_name=row.get("full_name"),
                    incident_title=title,
                    incident_message=msg[:500],
                    severity=_safe_int(incident_data.get("severity")),
                    distance_km=distance,
                )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Geofence alert error: {e}")


# ============================================================================
# AI MODERATION ENDPOINTS
# ============================================================================

@bp.route("/api/ai/analysis/<int:iid>", methods=["GET"])
@login_required
def api_get_ai_analysis(iid):
    """Obtiene el análisis IA de un incidente."""
    try:
        row = db.session.execute(text("""
            SELECT id, incident_id, decision, confidence, reason, flags,
                   alert_level, model_used, tokens_used, latency_ms, created_at
            FROM ai_analyses
            WHERE incident_id = :iid
            ORDER BY created_at DESC LIMIT 1
        """), {"iid": iid}).mappings().first()

        if not row:
            return jsonify({"error": "Sin análisis IA para este incidente"}), 404

        from ..models import Incident
        import json as json_mod

        incident = Incident.query.get(iid)
        reporter_profile = build_reporter_profiles(db.session, [incident.user_id]).get(incident.user_id, {}) if incident else {}
        ai_analysis = {
            "decision": row["decision"],
            "confidence": float(row["confidence"]),
            "reason": row["reason"],
            "flags": json_mod.loads(row["flags"]) if row["flags"] else [],
            "alert_level": row["alert_level"],
        }
        return jsonify({
            "id": row["id"],
            "incident_id": row["incident_id"],
            **ai_analysis,
            "model_used": row["model_used"],
            "tokens_used": row["tokens_used"],
            "latency_ms": row["latency_ms"],
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "reporter_profile": reporter_profile,
            "explainability": explain_analysis(
                {"incident_id": iid, "status": incident.status if incident else None},
                ai_analysis,
                reporter_profile,
            ),
            "evidence": list_incident_evidence(current_app.config["INCIDENT_MEDIA_ROOT"], iid),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ai/moderate/<int:iid>", methods=["POST"])
@login_required
def api_moderate_incident(iid):
    """Ejecuta moderación IA manual sobre un incidente (admin/moderator)."""
    from ..models import User, Incident
    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() not in ("admin", "moderator"):
        return jsonify({"error": "No autorizado"}), 403

    incident = Incident.query.get(iid)
    if not incident:
        return jsonify({"error": "Incidente no encontrado"}), 404

    try:
        from .. import ai_service
        svc = ai_service
        if not svc.is_available:
            return jsonify({"error": "Servicio IA no disponible. Verificar OPENAI_API_KEY."}), 503

        cat_name = incident.category.name if incident.category else "General"
        result = svc.moderate_incident(
            incident.title, incident.description, cat_name,
            float(incident.lat), float(incident.lng)
        )
        svc.save_analysis(iid, result)

        return jsonify({
            "success": True,
            "analysis": {
                "decision": result.get("decision"),
                "confidence": result.get("confidence"),
                "reason": result.get("reason"),
                "flags": result.get("flags", []),
                "alert_level": result.get("alert_level")
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ai/stats", methods=["GET"])
@login_required
def api_ai_stats():
    """Estadísticas del servicio IA (admin)."""
    from ..models import User
    user = User.query.get(session.get("user_id"))
    if not user or user.role_name() != "admin":
        return jsonify({"error": "No autorizado"}), 403

    try:
        stats = db.session.execute(text("""
            SELECT
                COUNT(*) as total_analyses,
                SUM(tokens_used) as total_tokens,
                AVG(latency_ms) as avg_latency,
                SUM(CASE WHEN decision='approved' THEN 1 ELSE 0 END) as approved,
                SUM(CASE WHEN decision='review' THEN 1 ELSE 0 END) as review,
                SUM(CASE WHEN decision='rejected' THEN 1 ELSE 0 END) as rejected
            FROM ai_analyses
        """)).mappings().first()

        from .. import ai_service
        return jsonify({
            "total_analyses": int(stats["total_analyses"] or 0),
            "total_tokens": int(stats["total_tokens"] or 0),
            "avg_latency_ms": round(float(stats["avg_latency"] or 0), 1),
            "approved": int(stats["approved"] or 0),
            "review": int(stats["review"] or 0),
            "rejected": int(stats["rejected"] or 0),
            "service_available": ai_service.is_available,
            "session_tokens": ai_service.total_tokens_used
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================================
# AI CHAT — Consulta ciudadana de zona
# ============================================================================

@bp.route("/api/ai/chat", methods=["POST"])
@login_required
def api_ai_chat():
    """Consulta IA avanzada con agente de analisis de datos."""
    data = request.get_json(silent=True) or {}
    history = _sanitize_chat_history(data.get("history") or [])
    try:
        message = _clean_text(data.get("message"), "Mensaje", MAX_CHAT_MESSAGE_LEN, min_len=2)
    except ValueError:
        return jsonify({"error": "Mensaje requerido"}), 400
    lat = data.get("lat")
    lng = data.get("lng")
    radius_km = data.get("radius_km") or 5
    place_name = re.sub(r"\s+", " ", str(data.get("place_name") or "")).strip()[:160]

    if lat is None or lng is None:
        lat, lng = 8.983, -79.517

    try:
        lat, lng = float(lat), float(lng)
        radius_km = _normalize_radius(radius_km, min_value=0.5, max_value=20)
    except (ValueError, TypeError):
        return jsonify({"error": "Coordenadas invalidas"}), 400
    try:
        _validate_panama_coords(lat, lng)
    except ValueError:
        return jsonify({"error": "Solo se admiten consultas dentro de Panama"}), 400

    try:
        rows = db.session.execute(text("""
            SELECT i.id, i.title, i.description, i.severity, i.status, i.created_at,
                   i.lat, i.lng, ic.name AS category_name
            FROM incidents i
            JOIN incident_categories ic ON i.category_id = ic.id
            WHERE i.created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
              AND (i.expires_at IS NULL OR i.expires_at >= NOW())
            ORDER BY i.created_at DESC
            LIMIT 250
        """
        )).mappings().all()

        incidents_list = []
        for r in rows:
            incidents_list.append({
                "id": r["id"], "title": r["title"], "description": r["description"],
                "severity": int(r["severity"]), "status": r["status"],
                "created_at": str(r["created_at"]), "lat": float(r["lat"]),
                "lng": float(r["lng"]), "category_name": r["category_name"]
            })

        from .. import ai_service as ai_svc_mod
        from ..ai_agents import DataAnalystAgent
        analyst = DataAnalystAgent(ai_svc_mod)
        result = analyst.answer_question(message, lat, lng, incidents_list, context={
            "radius_km": radius_km,
            "place_name": place_name,
            "history": history
        })
        result['assistant_mode'] = _get_plan_context(session.get("user_id"))["assistant_mode"]
        result['place_name'] = result.get('place_name') or place_name or f"{lat:.4f}, {lng:.4f}"
        result['radius_km'] = radius_km
        result['history_used'] = len(history)

        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ai/zone-analysis", methods=["POST"])
@login_required
def api_zone_analysis():
    """Analisis completo de zona con agente de datos."""
    data = request.get_json(silent=True) or {}
    lat = data.get("lat")
    lng = data.get("lng")
    if lat is None or lng is None:
        return jsonify({"error": "Coordenadas requeridas"}), 400

    try:
        lat, lng = float(lat), float(lng)
        radius = _normalize_radius(data.get("radius_km", 3), min_value=0.5, max_value=20)
    except (ValueError, TypeError):
        return jsonify({"error": "Coordenadas invalidas"}), 400
    try:
        _validate_panama_coords(lat, lng)
    except ValueError:
        return jsonify({"error": "Solo se admiten zonas dentro de Panama"}), 400

    try:
        rows = db.session.execute(text("""
            SELECT i.id, i.title, i.description, i.severity, i.status, i.created_at,
                   i.lat, i.lng, ic.name AS category_name
            FROM incidents i
            JOIN incident_categories ic ON i.category_id = ic.id
            WHERE i.created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
              AND (i.expires_at IS NULL OR i.expires_at >= NOW())
            ORDER BY i.created_at DESC LIMIT 200
        """)).mappings().all()

        incidents_list = [{
            "id": r["id"], "title": r["title"], "description": r["description"],
            "severity": int(r["severity"]), "status": r["status"],
            "created_at": str(r["created_at"]), "lat": float(r["lat"]),
            "lng": float(r["lng"]), "category_name": r["category_name"]
        } for r in rows]

        from .. import ai_service as ai_svc_mod
        from ..ai_agents import DataAnalystAgent
        analyst = DataAnalystAgent(ai_svc_mod)
        result = analyst.analyze_zone(lat, lng, incidents_list, radius)

        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ai/fake-check", methods=["POST"])
@login_required
def api_fake_check():
    """Verifica si un reporte es potencialmente falso usando el agente detector."""
    data = request.get_json(silent=True) or {}

    try:
        # Get user's recent history
        user_history = db.session.execute(text("""
            SELECT id, title, description, lat, lng, severity, created_at
            FROM incidents WHERE user_id = :uid
            ORDER BY created_at DESC LIMIT 20
        """), {"uid": session["user_id"]}).mappings().all()

        user_hist = [{
            "id": r["id"], "title": r["title"], "lat": float(r["lat"]),
            "lng": float(r["lng"]), "severity": int(r["severity"]),
            "created_at": str(r["created_at"])
        } for r in user_history]

        # Get recent nearby incidents
        lat, lng = float(data.get("lat", 0)), float(data.get("lng", 0))
        nearby = db.session.execute(text("""
            SELECT i.id, i.title, i.description, i.category_id, i.lat, i.lng,
                   ic.name AS category_name
            FROM incidents i
            JOIN incident_categories ic ON i.category_id = ic.id
            WHERE i.created_at >= DATE_SUB(NOW(), INTERVAL 24 HOUR)
            AND ST_Distance_Sphere(POINT(i.lng, i.lat), POINT(:lng, :lat)) / 1000 <= 2
            ORDER BY i.created_at DESC LIMIT 20
        """), {"lat": lat, "lng": lng}).mappings().all()

        nearby_list = [dict(r) for r in nearby]
        for n in nearby_list:
            for k, v in n.items():
                if hasattr(v, '__float__') and not isinstance(v, (int, bool)):
                    n[k] = float(v)

        from .. import ai_service as ai_svc_mod
        from ..ai_agents import FakePingDetector
        detector = FakePingDetector(ai_svc_mod)
        result = detector.analyze(data, user_hist, nearby_list)

        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ai/dashboard-insights", methods=["GET"])
@login_required
def api_dashboard_insights():
    """Genera insights IA para el dashboard."""
    try:
        rows = db.session.execute(text("""
            SELECT i.id, i.title, i.severity, i.status, i.created_at,
                   i.lat, i.lng, ic.name AS category_name
            FROM incidents i
            JOIN incident_categories ic ON i.category_id = ic.id
            WHERE i.created_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)
              AND (i.expires_at IS NULL OR i.expires_at >= NOW())
            ORDER BY i.created_at DESC LIMIT 100
        """)).mappings().all()

        incidents_list = [{
            "id": r["id"], "title": r["title"],
            "severity": int(r["severity"]), "status": r["status"],
            "created_at": str(r["created_at"]),
            "lat": float(r["lat"]), "lng": float(r["lng"]),
            "category_name": r["category_name"]
        } for r in rows]

        # Panama City center for general analysis
        from .. import ai_service as ai_svc_mod
        from ..ai_agents import DataAnalystAgent
        analyst = DataAnalystAgent(ai_svc_mod)
        result = analyst.analyze_zone(8.983, -79.517, incidents_list, radius_km=20)

        return jsonify({
            "insights": result.get("insights", []),
            "predictions": result.get("predictions", []),
            "risk_level": result.get("risk_level", "low"),
            "trend": result.get("trend", {}),
            "statistics": result.get("statistics", {})
        })
    except Exception as e:
        return jsonify({"insights": [], "predictions": [], "error": str(e)})


# ============================================================================
# UPGRADE DE PLAN (simulado — sin pasarela de pago real)
# ============================================================================

@bp.route("/api/plans/upgrade", methods=["POST"])
@login_required
def api_upgrade_plan():
    """Upgrade de plan simulado. Requiere método de pago registrado."""
    from ..models import Plan, Subscription, PaymentMethod
    from datetime import datetime, timedelta

    b = request.get_json(silent=True) or {}
    plan_id = b.get("plan_id")

    if not plan_id:
        return jsonify({"error": "plan_id requerido"}), 400

    plan = Plan.query.get(int(plan_id))
    if not plan:
        return jsonify({"error": "Plan no encontrado"}), 404

    # Verificar que tenga método de pago si es plan de pago
    if float(plan.price_monthly_usd) > 0:
        pm = PaymentMethod.query.filter_by(user_id=session["user_id"]).first()
        if not pm:
            return jsonify({"error": "Registra un método de pago primero"}), 400

    try:
        sub = Subscription.query.filter_by(user_id=session["user_id"]).first()
        if sub:
            sub.plan_id = plan.id
            sub.started_at = datetime.utcnow()
            sub.expires_at = datetime.utcnow() + timedelta(days=30)
            sub.status = "active"
        else:
            sub = Subscription(
                user_id=session["user_id"], plan_id=plan.id,
                started_at=datetime.utcnow(),
                expires_at=datetime.utcnow() + timedelta(days=30),
                status="active"
            )
            db.session.add(sub)
        db.session.commit()

        session["plan"] = plan.name
        return jsonify({"success": True, "plan": plan.name})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


# ============================================================================
# PLANS (público)
# ============================================================================

@bp.route("/api/plans", methods=["GET"])
def api_plans():
    """Lista planes disponibles."""
    from ..models import Plan
    plans = Plan.query.order_by(Plan.price_monthly_usd).all()
    return jsonify([{
        "id": p.id,
        "name": p.name,
        "price_monthly_usd": float(p.price_monthly_usd),
        "max_alerts_per_day": p.max_alerts_per_day,
        "max_geo_fences": p.max_geo_fences,
        "api_access": p.api_access,
        "priority_support": p.priority_support,
        "features": p.features_json
    } for p in plans])
